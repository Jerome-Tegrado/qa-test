from pathlib import Path
import math
import re
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
TC_OUTPUT_DIR = OUTPUT_DIR / "test-cases"

TC_CUSTOMIZE_FILE = INPUT_DIR / "customize" / "test-case.yaml"
LEGACY_TC_FORMAT_FILE = INPUT_DIR / "format" / "tc-format.yaml"
TC_DATA_FILE = INPUT_DIR / "tc" / "tc.yaml"
TP_DATA_FILE = INPUT_DIR / "tp" / "tp.yaml"

DEFAULT_TC_CUSTOMIZATION = {
    "defaults": {
        "workbook": {
            "sheet_name": "Sheet1",
            "start_row": 1,
            "start_column": 1,
        },
        "columns": [
            {"key": "title", "header": "Test Case Title", "column": "A", "width": 40},
            {"key": "description", "header": "Test Description", "column": "B", "width": 42},
            {"key": "preconditions", "header": "Preconditions", "column": "C", "width": 34},
            {"key": "steps", "header": "Test Steps", "column": "D", "width": 34},
            {"key": "expected_results", "header": "Expected Results", "column": "E", "width": 36},
            {"key": "actual_results", "header": "Actual Results", "column": "F", "width": 36},
            {"key": "status", "header": "Status", "column": "G", "width": 15},
            {"key": "test_evidence", "header": "Test Evidence", "column": "H", "width": 24},
        ],
        "row_values": {
            "actual_results": "",
            "status": "",
            "test_evidence": "",
        },
    },
    "style": {
        "header": {
            "font_name": "Calibri",
            "font_size": 11,
            "bold": True,
            "font_color": "FFFFFF",
            "fill_color": "76933C",
            "horizontal_alignment": "center",
            "vertical_alignment": "center",
            "wrap_text": True,
        },
        "body": {
            "font_name": "Calibri",
            "font_size": 10,
            "bold": False,
            "font_color": "000000",
            "horizontal_alignment": "left",
            "vertical_alignment": "top",
            "wrap_text": True,
            "border": {"style": "thin", "color": "D9D9D9"},
        },
    },
    "layout": {
        "row_style": {"header_height": 22, "body_height": 60},
        "page_setup": {
            "orientation": "landscape",
            "fit_to_width": 1,
            "fit_to_height": 0,
            "margins": {
                "left": 0.25,
                "right": 0.25,
                "top": 0.5,
                "bottom": 0.5,
                "header": 0.3,
                "footer": 0.3,
            },
        },
    },
    "behavior": {
        "generation_rules": {
            "freeze_header_row": True,
            "status_dropdown_options": ["Passed", "Failed"],
            "blank_rows_when_no_data": 6,
        }
    },
}


def safe_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*]', "-", name or "")
    name = re.sub(r"\s+", " ", name).strip()
    return name[:80] or "Title"


def load_yaml(file_path: Path) -> dict:
    with open(file_path, "r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {}


def deep_merge(base, override):
    if isinstance(base, dict) and isinstance(override, dict):
        merged = dict(base)
        for key, value in override.items():
            if key in merged:
                merged[key] = deep_merge(merged[key], value)
            else:
                merged[key] = value
        return merged
    return override


def normalize_tc_customization_schema(raw_cfg: dict) -> dict:
    if not raw_cfg:
        return {}

    if "defaults" in raw_cfg or "style" in raw_cfg or "layout" in raw_cfg or "behavior" in raw_cfg:
        return raw_cfg

    # Backward compatibility with legacy top-level tc-format.yaml schema.
    return {
        "defaults": {
            "workbook": raw_cfg.get("workbook", {}),
            "columns": raw_cfg.get("columns", []),
            "row_values": raw_cfg.get("row_values", {}),
        },
        "style": {
            "header": raw_cfg.get("header_style", {}),
            "body": raw_cfg.get("body_style", {}),
        },
        "layout": {
            "row_style": raw_cfg.get("row_style", {}),
            "page_setup": raw_cfg.get("page_setup", {}),
        },
        "behavior": {"generation_rules": raw_cfg.get("generation_rules", {})},
    }


def load_tc_customization() -> dict:
    if TC_CUSTOMIZE_FILE.exists():
        raw_cfg = load_yaml(TC_CUSTOMIZE_FILE)
    elif LEGACY_TC_FORMAT_FILE.exists():
        raw_cfg = load_yaml(LEGACY_TC_FORMAT_FILE)
    else:
        raw_cfg = {}

    normalized_cfg = normalize_tc_customization_schema(raw_cfg)
    merged_cfg = deep_merge(DEFAULT_TC_CUSTOMIZATION, normalized_cfg)

    return {
        "workbook": merged_cfg.get("defaults", {}).get("workbook", {}),
        "columns": merged_cfg.get("defaults", {}).get("columns", []),
        "row_values": merged_cfg.get("defaults", {}).get("row_values", {}),
        "header_style": merged_cfg.get("style", {}).get("header", {}),
        "body_style": merged_cfg.get("style", {}).get("body", {}),
        "row_style": merged_cfg.get("layout", {}).get("row_style", {}),
        "page_setup": merged_cfg.get("layout", {}).get("page_setup", {}),
        "generation_rules": merged_cfg.get("behavior", {}).get("generation_rules", {}),
    }


def coalesce(value, default):
    if value is None:
        return default
    if isinstance(value, str) and not value.strip():
        return default
    return value


def build_font(style: dict) -> Font:
    return Font(
        name=style.get("font_name", "Calibri"),
        size=style.get("font_size", 11),
        bold=style.get("bold", False),
        color=style.get("font_color", "000000"),
    )


def build_alignment(style: dict) -> Alignment:
    return Alignment(
        horizontal=style.get("horizontal_alignment", "left"),
        vertical=style.get("vertical_alignment", "top"),
        wrap_text=style.get("wrap_text", True),
    )


def build_border(style: dict) -> Border:
    side_style = style.get("style", "thin")
    color = style.get("color", "D9D9D9")
    side = Side(border_style=side_style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def to_numbered_list(items: list[str]) -> str:
    if not items:
        return ""
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items, start=1))


def to_bullet_list(items: list[str]) -> str:
    if not items:
        return ""
    return "\n".join(f"\u2022 {item}" for item in items)


def to_single_bullet(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return ""
    return f"\u2022 {text}"


def normalize_items(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    if isinstance(value, str):
        lines: list[str] = []
        for raw in value.splitlines():
            line = raw.strip()
            if not line:
                continue
            line = re.sub(r"^(?:(?:[-*]|\u2022)+\s*|\d+[\.)]\s*)", "", line).strip()
            if line:
                lines.append(line)
        return lines
    text = str(value).strip()
    return [text] if text else []


def validate_test_case_structure(test_case: dict, case_index: int) -> None:
    description = (test_case.get("description") or "").strip()
    precondition_value = (
        test_case.get("precondition")
        if test_case.get("precondition") is not None
        else test_case.get("preconditions")
    )
    preconditions = normalize_items(precondition_value)
    steps = normalize_items(test_case.get("steps"))
    expected_results = normalize_items(test_case.get("expected_results"))

    if not description:
        raise ValueError(f"Test case #{case_index}: description is required.")

    if "\n" in description:
        raise ValueError(
            f"Test case #{case_index}: description must be 1 sentence only and must not contain line breaks."
        )

    if not preconditions:
        raise ValueError(f"Test case #{case_index}: precondition is required.")

    if len(steps) == 0:
        raise ValueError(f"Test case #{case_index}: steps must contain at least 1 item.")

    if len(expected_results) == 0:
        raise ValueError(f"Test case #{case_index}: expected_results must contain at least 1 item.")


def set_page_setup(ws, page_setup: dict) -> None:
    ws.page_setup.orientation = page_setup.get("orientation", "landscape")
    ws.page_setup.fitToWidth = page_setup.get("fit_to_width", 1)
    ws.page_setup.fitToHeight = page_setup.get("fit_to_height", 0)

    margins = page_setup.get("margins", {})
    ws.page_margins.left = margins.get("left", 0.25)
    ws.page_margins.right = margins.get("right", 0.25)
    ws.page_margins.top = margins.get("top", 0.5)
    ws.page_margins.bottom = margins.get("bottom", 0.5)
    ws.page_margins.header = margins.get("header", 0.3)
    ws.page_margins.footer = margins.get("footer", 0.3)


def estimate_row_height(ws, row_index: int, columns: list[dict], minimum_height: float) -> float:
    # Approximate wrapped line count per cell using column width as character capacity.
    max_lines = 1
    for col_cfg in columns:
        col_letter = col_cfg["column"]
        value = ws[f"{col_letter}{row_index}"].value
        if value is None:
            continue

        text = str(value)
        width = ws.column_dimensions[col_letter].width or col_cfg.get("width", 20)
        chars_per_line = max(int(width * 0.95), 8)

        lines_for_cell = 0
        for raw_line in text.split("\n"):
            line = raw_line or " "
            lines_for_cell += max(1, math.ceil(len(line) / chars_per_line))

        max_lines = max(max_lines, lines_for_cell)

    # Calibri 11 with wrapped text is typically readable around 14-15pt per line.
    estimated = max_lines * 14.5
    return max(minimum_height, estimated)


def add_status_dropdown(
    ws,
    start_row: int,
    end_row: int,
    status_col_letter: str,
    options: list[str]
) -> None:
    formula = '"' + ",".join(options) + '"'
    dropdown = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True
    )
    dropdown.prompt = "Select status"
    dropdown.promptTitle = "Test Status"
    dropdown.error = f"Choose only from: {', '.join(options)}"
    dropdown.errorTitle = "Invalid Status"

    ws.add_data_validation(dropdown)
    dropdown.add(f"{status_col_letter}{start_row}:{status_col_letter}{end_row}")


def main() -> None:
    TC_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    fmt = load_tc_customization()
    data = load_yaml(TC_DATA_FILE)
    tp_data = load_yaml(TP_DATA_FILE) if TP_DATA_FILE.exists() else {}

    workbook_cfg = fmt.get("workbook", {})
    columns = fmt.get("columns", [])
    header_style = fmt.get("header_style", {})
    body_style = fmt.get("body_style", {})
    row_style = fmt.get("row_style", {})
    page_setup = fmt.get("page_setup", {})
    rules = fmt.get("generation_rules", {})
    row_values = fmt.get("row_values", {})

    sheet_name = data.get("sheet_name") or workbook_cfg.get("sheet_name", "Sheet1")
    test_cases = data.get("test_cases") or data.get("rows") or []
    default_title = (
        tp_data.get("tp_title")
        or
        data.get("tc_title")
        or data.get("test_case_title")
        or data.get("title")
        or (test_cases[0].get("title") if test_cases else "")
        or "Title"
    )
    output_file_name = data.get("output_file_name") or f"Test Case_{safe_filename(default_title)}.xlsx"

    header_row = workbook_cfg.get("start_row", 1)
    start_column = workbook_cfg.get("start_column", 1)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = build_font(header_style)
    body_font = build_font(body_style)
    header_alignment = build_alignment(header_style)
    body_alignment = build_alignment(body_style)
    body_border = build_border(body_style.get("border", {}))

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=header_style.get("fill_color", "76933C")
    )

    # Write headers
    for index, column_cfg in enumerate(columns, start=start_column):
        cell = ws.cell(row=header_row, column=index)
        cell.value = column_cfg["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

        column_letter = column_cfg["column"]
        ws.column_dimensions[column_letter].width = column_cfg.get("width", 20)

    ws.row_dimensions[header_row].height = row_style.get("header_height", 22)

    # Write test case rows
    current_row = header_row + 1
    for case_index, test_case in enumerate(test_cases, start=1):
        validate_test_case_structure(test_case, case_index)

        row_map = {
            "title": test_case.get("title", "").strip(),
            "description": test_case.get("description", "").strip(),
            "preconditions": to_bullet_list(
                normalize_items(
                    test_case.get("precondition")
                    if test_case.get("precondition") is not None
                    else test_case.get("preconditions")
                )
            ),
            "steps": to_numbered_list(normalize_items(test_case.get("steps"))),
            "expected_results": to_bullet_list(normalize_items(test_case.get("expected_results"))),
            "actual_results": coalesce(test_case.get("actual_results"), row_values.get("actual_results", "")),
            "status": coalesce(test_case.get("status"), row_values.get("status", "")),
            "test_evidence": coalesce(test_case.get("test_evidence"), row_values.get("test_evidence", "")),
        }

        for index, column_cfg in enumerate(columns, start=start_column):
            key = column_cfg["key"]
            cell = ws.cell(row=current_row, column=index)
            cell.value = row_map.get(key, "")
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = body_border

        ws.row_dimensions[current_row].height = estimate_row_height(
            ws=ws,
            row_index=current_row,
            columns=columns,
            minimum_height=row_style.get("body_height", 60),
        )
        current_row += 1

    # Create blank rows when there is no data
    if not test_cases:
        blank_rows = rules.get("blank_rows_when_no_data", 6)
        for row in range(header_row + 1, header_row + 1 + blank_rows):
            for index, column_cfg in enumerate(columns, start=start_column):
                cell = ws.cell(row=row, column=index)
                cell.value = ""
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = body_border
            ws.row_dimensions[row].height = row_style.get("body_height", 60)
        current_row = header_row + 1 + blank_rows

    # Freeze header row
    if rules.get("freeze_header_row", True):
        ws.freeze_panes = f"A{header_row + 1}"

    # Add status dropdown in Status column only
    last_row = max(current_row - 1, header_row + 20)
    status_column_letter = None
    for column_cfg in columns:
        if column_cfg["key"] == "status":
            status_column_letter = column_cfg["column"]
            break

    if status_column_letter:
        add_status_dropdown(
            ws=ws,
            start_row=header_row + 1,
            end_row=last_row,
            status_col_letter=status_column_letter,
            options=rules.get("status_dropdown_options", ["Passed", "Failed"]),
        )

    set_page_setup(ws, page_setup)

    output_path = TC_OUTPUT_DIR / Path(output_file_name).name
    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    main()
